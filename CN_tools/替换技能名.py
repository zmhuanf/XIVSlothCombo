
# import openpyxl
# import re

# # 读取 xlsx 文件
# workbook = openpyxl.load_workbook('技能名data1.xlsx')
# worksheet = workbook.active

# # 构建替换字典
# replace_dict = {}
# for row in worksheet.iter_rows(values_only=True):
#     replace_dict[row[0]] = row[1]

# # 读取 txt 文件
# with open('CustomComboPreset.txt', 'r',encoding="utf8") as f:
#     content = f.read()

# # 替换文本
# for key, value in replace_dict.items():
#     print(f"Replacing '{key}' with '{value}'")
#     if key is not None and value is not None:
#         # 使用正则表达式匹配被""包裹的语句
#         pattern = r'"([^"]*?(' + key + r')[^"]*)"'  # 在匹配表达式中使用非贪婪模式匹配最小的语句
#         content = re.sub(pattern, lambda match: match.group(0).replace(match.group(2), value), content)
#     else:
#         print(f"Skipping {key} -> {value} because it contains None")
#         continue

# # 输出替换后的 txt 文件
# with open('output.txt', 'w',encoding="utf8") as f:
#     f.write(content)
import openpyxl
import re

# 读取xlsx表格
wb = openpyxl.load_workbook('技能名data1.xlsx')
ws = wb.active

# 将第一列作为替换词，第二列作为被替换词
replace_dict = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    key = row[0]
    value = row[1]
    if key is not None and value is not None:
        replace_dict[key] = value

# 读取txt文件
with open('CustomComboPreset.txt', 'r', encoding='utf-8') as f:
    content = f.read()

# 匹配被""包裹的子语句并替换
pattern = r'"(?:\\.|[^"\\])*"'  # 正则表达式匹配被""包裹的子语句
matches = re.findall(pattern, content)
for match in matches:
    print("1")
    for key, value in replace_dict.items():
        if match.strip('"') == key:
            content = content.replace(match, '"' + value + '"')

# 输出替换后的txt文件
with open('modified.txt', 'w', encoding='utf-8') as f:
    f.write(content)
